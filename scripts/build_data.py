#!/usr/bin/env python3
"""
MyClinical Growth — data builder (one-time / on-demand).

Generates:
  data/directory.json          from the source directory spreadsheet
  data/opportunities.json       curated standing frameworks & routes (real)
  data/grants.json              curated grant programmes (real, verified)
  data/opportunities-live.json  empty shell — the poller fills this on deploy

The opportunities/grants here are real, verified sources. The continuous
feed of individual tender notices is added by scripts/poll.py once the site
is deployed (the poller needs open internet, which GitHub Actions provides).
"""

import json
import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
XLSX = ROOT.parent / "NHS Procurement & Grants Source Directory.xlsx"
NOW = dt.datetime.utcnow().isoformat() + "Z"


def build_directory():
    wb = load_workbook(XLSX, read_only=True)
    out = {"procurement": [], "grants": []}
    for sheet, key in (("Procurement", "procurement"), ("Grants & Funding", "grants")):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in rows[0]]
        for r in rows[1:]:
            rec = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
            out[key].append(rec)
    wb.close()
    payload = {"updated": NOW,
               "procurement_count": len(out["procurement"]),
               "grant_count": len(out["grants"]),
               **out}
    (DATA / "directory.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    print(f"directory.json — {len(out['procurement'])} procurement, {len(out['grants'])} grants")


STANDING_OPPORTUNITIES = [
    {
        "id": "std-sbs-ai-solutions",
        "title": "Healthcare AI Solutions Framework",
        "source": "NHS Shared Business Services",
        "source_url": "https://www.find-tender.service.gov.uk/Notice/001651-2025",
        "buyer": "NHS Shared Business Services",
        "type": "Framework", "category": "AI & analytics",
        "value": "£900,000,000", "published": "2026-05-13", "deadline": "2026-06-23",
        "status": "Open",
        "summary": "National framework for healthcare AI across eight lots — radiology and imaging, pathology and early detection, virtual and robotic health, predictive analytics, research and development, operational efficiency, advisory and specialised support, and combined solutions.",
        "means": "The procurement route for UK healthcare AI for the next four years. Lots 6 and 7 — operational efficiency and advisory — are the quieter opportunity. The 23 June deadline is tight: clinical evidence packs and DCB0129 documentation take weeks.",
        "tags": ["Framework", "AI", "Closes soon"],
    },
    {
        "id": "std-gp-it-futures",
        "title": "Tech Innovation Framework / GP IT Futures",
        "source": "NHS England Digital",
        "source_url": "https://digital.nhs.uk/services/digital-care-services-catalogue/tech-innovation-framework",
        "buyer": "NHS England", "type": "Framework", "category": "Clinical systems & EPR",
        "value": "", "published": "2026-04-01", "deadline": "",
        "status": "Rolling",
        "summary": "Assured core GP IT systems and digital products for primary care, bought via the Digital Care Services catalogue. Tech Innovation Framework is the successor to GP IT Futures.",
        "means": "Essential route for anything touching primary care. GP IT Futures continuity contracts run only to June 2026 — if primary care is a target market, getting onto TIF is the gating step.",
        "tags": ["Framework", "Primary care"],
    },
    {
        "id": "std-medtech-funding-mandate",
        "title": "MedTech Funding Mandate",
        "source": "NHS England / Accelerated Access Collaborative",
        "source_url": "https://www.england.nhs.uk/aac/what-we-do/how-can-the-aac-help-me/the-medtech-funding-mandate/",
        "buyer": "NHS England", "type": "Adoption route", "category": "Other digital health",
        "value": "", "published": "2026-04-01", "deadline": "",
        "status": "Ongoing",
        "summary": "Mandates NHS commissioners and providers to fund a defined list of NICE-approved cost-saving devices, diagnostics and digital products.",
        "means": "Not a tender — a route to guaranteed local funding. Eligibility: positive NICE guidance, net saving within three years, budget impact under £20m a year. Worth understanding early if a NICE pathway is in reach.",
        "tags": ["Adoption route", "NICE"],
    },
    {
        "id": "std-gcloud",
        "title": "G-Cloud (Digital Marketplace)",
        "source": "Crown Commercial Service",
        "source_url": "https://www.digitalmarketplace.service.gov.uk/",
        "buyer": "Crown Commercial Service", "type": "Framework", "category": "Infrastructure & cloud",
        "value": "", "published": "2026-01-01", "deadline": "",
        "status": "Rolling",
        "summary": "Cross-government framework for cloud hosting, software (SaaS) and digital support. The primary procurement route for software-based digital health products.",
        "means": "The most relevant framework for digital health SaaS. Being listed makes a product discoverable to every NHS buyer. Application windows open periodically, worth tracking.",
        "tags": ["Framework", "Cloud"],
    },
]

GRANTS = [
    {
        "id": "grant-sbri-healthcare",
        "title": "SBRI Healthcare — themed competitions",
        "source": "SBRI Healthcare",
        "source_url": "https://sbrihealthcare.co.uk/competitions",
        "type": "Grant programme", "category": "NHS-aligned innovation",
        "value": "Phase 1 £50k–£150k · Phase 2 up to £1m",
        "published": "2026-05-01", "deadline": "", "status": "Competitions in waves",
        "summary": "Themed competitions for innovations addressing unmet NHS need — open to digital, AI and device innovations. Includes an Investment Readiness programme.",
        "means": "The flagship NHS-aligned non-dilutive funder. As much a credibility marker as a cash injection — an SBRI award signals NHS validation to later investors.",
        "tags": ["Grant", "Non-dilutive"],
    },
    {
        "id": "grant-nihr-i4i",
        "title": "NIHR Invention for Innovation (i4i) — Product Development Awards",
        "source": "National Institute for Health and Care Research",
        "source_url": "https://www.nihr.ac.uk/research-funding/funding-programmes/invention-for-innovation",
        "type": "Grant programme", "category": "Translational funding",
        "value": "SMEs eligible for up to 100% funding",
        "published": "2026-05-01", "deadline": "", "status": "Periodic calls",
        "summary": "Translational funding for medical devices, IVDs and digital health technologies bridging proof-of-concept to real-world clinical adoption. Two-stage application.",
        "means": "Best fit for a product with early clinical evidence that needs the push to adoption-readiness. The two-stage process means an outline can be tested cheaply before committing to a full bid.",
        "tags": ["Grant", "Non-dilutive"],
    },
    {
        "id": "grant-nihr-eme",
        "title": "NIHR Efficacy and Mechanism Evaluation (EME) Programme",
        "source": "National Institute for Health and Care Research",
        "source_url": "https://www.nihr.ac.uk/research-funding/funding-programmes/efficacy-and-mechanism-evaluation",
        "type": "Grant programme", "category": "Clinical evidence",
        "value": "Varies", "published": "2026-05-01", "deadline": "2026-08-05",
        "status": "Open — outline stage",
        "summary": "Funds evaluation of efficacy and the mechanisms behind health interventions, including digital and technology-enabled interventions.",
        "means": "Relevant where the gap is hard clinical evidence rather than product build. Outline applications close 5 August 2026.",
        "tags": ["Grant", "Closes soon"],
    },
    {
        "id": "grant-biomedical-catalyst",
        "title": "Innovate UK — Biomedical Catalyst",
        "source": "Innovate UK (UKRI)",
        "source_url": "https://www.ukri.org/councils/innovate-uk/",
        "type": "Grant programme", "category": "R&D funding",
        "value": "Up to ~£2m for large projects",
        "published": "2026-05-01", "deadline": "", "status": "Rolling competitions",
        "summary": "Supports SMEs developing innovative solutions to health and healthcare challenges, from small projects through to large collaborative R&D.",
        "means": "The broadest UK non-dilutive R&D funder for healthtech. Pairs well with SBRI/NIHR — different stages of the same journey.",
        "tags": ["Grant", "R&D"],
    },
    {
        "id": "grant-eic-accelerator",
        "title": "EIC Accelerator",
        "source": "Horizon Europe / European Innovation Council",
        "source_url": "https://eic.ec.europa.eu/eic-funding-opportunities/eic-accelerator_en",
        "type": "International grant", "category": "Deep-tech funding",
        "value": "Up to ~€2.5m grant + equity",
        "published": "2026-05-01", "deadline": "2026-07-08", "status": "Open — next cut-off",
        "summary": "Grant plus equity for deep-tech SMEs and startups. UK organisations are eligible following association to Horizon Europe.",
        "means": "Six fixed cut-offs a year — predictable enough to plan a submission around. Next cut-off 8 July 2026; further cut-offs 2 September and 4 November.",
        "tags": ["Grant", "International", "Closes soon"],
    },
    {
        "id": "grant-horizon-health",
        "title": "Horizon Europe — Health Cluster (Cluster 1)",
        "source": "European Commission",
        "source_url": "https://research-and-innovation.ec.europa.eu/funding/funding-opportunities/funding-programmes-and-open-calls/horizon-europe/cluster-1-health_en",
        "type": "International grant", "category": "Collaborative R&I",
        "value": "Up to ~€90m per call",
        "published": "2026-05-01", "deadline": "", "status": "2026–27 work programme open",
        "summary": "Collaborative health research and innovation funding. UK organisations can apply as part of a consortium of at least three countries, one an EU member state.",
        "means": "Big money, but consortium-bound — best approached as a partner in a larger bid rather than a solo route. The 2026–27 work programme is published; specific calls have fixed deadlines.",
        "tags": ["Grant", "International"],
    },
]


def build_curated():
    (DATA / "opportunities.json").write_text(json.dumps(
        {"updated": NOW, "count": len(STANDING_OPPORTUNITIES),
         "opportunities": STANDING_OPPORTUNITIES}, indent=2, ensure_ascii=False))
    print(f"opportunities.json — {len(STANDING_OPPORTUNITIES)} standing items")

    (DATA / "grants.json").write_text(json.dumps(
        {"updated": NOW, "count": len(GRANTS), "grants": GRANTS},
        indent=2, ensure_ascii=False))
    print(f"grants.json — {len(GRANTS)} grant programmes")

    live = DATA / "opportunities-live.json"
    if not live.exists():
        live.write_text(json.dumps(
            {"updated": None, "count": 0, "opportunities": [],
             "note": "Populated by scripts/poll.py once deployed — the poller needs open internet (GitHub Actions provides it)."},
            indent=2, ensure_ascii=False))
        print("opportunities-live.json — empty shell created")


if __name__ == "__main__":
    DATA.mkdir(exist_ok=True)
    build_directory()
    build_curated()
    print("Done.")
